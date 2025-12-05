"""
Excel Report Utility Functions

Shared utilities for generating professional Excel reports across the KPI pipeline.
Provides consistent formatting, styling, and helper functions for report generation.

This module eliminates duplication between generate_reports.py and generate_pm_reports.py
by extracting common formatting logic into reusable functions.

Author: KPI Pipeline Project
Date: 2025-12-03
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional


# ============================================================================
# COLOR CONSTANTS
# ============================================================================

# Status Colors (traffic light system)
COLOR_GREEN = "C6EFCE"           # Light green - success/meeting targets
COLOR_LIGHT_GREEN = "90EE90"     # Pale green - excellent performance
COLOR_PALE_GREEN = "B4F8C8"      # Pale green - good performance
COLOR_YELLOW = "FFEB9C"          # Light yellow - warning/at risk
COLOR_ORANGE = "FFD580"          # Light orange - needs improvement
COLOR_RED = "FFC7CE"             # Light red - critical/failure
COLOR_LIGHT_RED = "FFB6C1"       # Light pink - poor performance

# Header Colors
COLOR_HEADER = "4472C4"          # Professional blue for headers
COLOR_SUBHEADER = "70AD47"       # Green for subheaders
COLOR_WHITE = "FFFFFF"           # White text

# Border Colors
COLOR_BORDER = "D0D0D0"          # Light gray borders

# Neutral Colors
COLOR_GRAY = "CCCCCC"            # Gray for unknown/neutral


# ============================================================================
# STATUS MAPPING FUNCTIONS
# ============================================================================

def get_status_color(status: str, color_scheme: str = 'default') -> str:
    """
    Get the fill color for a given status string.

    Supports multiple color schemes to accommodate different status
    nomenclatures used across KPI and OKR reports.

    Args:
        status: Status string (case-insensitive)
        color_scheme: Color scheme to use ('default', 'kpi', 'traffic')

    Returns:
        Hex color code for the status

    Examples:
        >>> get_status_color('Met')
        'C6EFCE'
        >>> get_status_color('CRITICAL')
        'FFC7CE'
        >>> get_status_color('GREEN', color_scheme='traffic')
        'C6EFCE'
    """
    status_upper = status.upper()

    # Traffic light system (GREEN/YELLOW/RED)
    if color_scheme == 'traffic':
        color_map = {
            'GREEN': COLOR_GREEN,
            'YELLOW': COLOR_YELLOW,
            'RED': COLOR_RED
        }
        return color_map.get(status_upper, COLOR_BORDER)

    # Comprehensive status mapping (default)
    color_map = {
        # Success statuses
        'MET': COLOR_GREEN,
        'PASS': COLOR_LIGHT_GREEN,
        'EXCELLENT': COLOR_LIGHT_GREEN,
        'GOOD': COLOR_PALE_GREEN,
        'ON TRACK': COLOR_GREEN,
        'GREEN': COLOR_GREEN,

        # Warning statuses
        'WARNING': COLOR_ORANGE,
        'AT RISK': COLOR_YELLOW,
        'NEEDS IMPROVEMENT': COLOR_ORANGE,
        'YELLOW': COLOR_YELLOW,

        # Critical statuses
        'CRITICAL': COLOR_RED,
        'FAIL': COLOR_LIGHT_RED,
        'POOR': COLOR_LIGHT_RED,
        'OFF TRACK': COLOR_RED,
        'RED': COLOR_RED,
    }

    return color_map.get(status_upper, COLOR_GRAY)


def get_status_emoji(status: str) -> str:
    """
    Get an emoji indicator for a given status string.

    Args:
        status: Status string (case-insensitive)

    Returns:
        Emoji string for the status

    Examples:
        >>> get_status_emoji('Met')
        '🟢'
        >>> get_status_emoji('CRITICAL')
        '🔴'
        >>> get_status_emoji('WARNING')
        '🟡'
    """
    status_upper = status.upper()

    emoji_map = {
        # Success emojis
        'MET': '🟢',
        'PASS': '🟢',
        'EXCELLENT': '🟢',
        'GOOD': '🟢',
        'ON TRACK': '🟢',
        'GREEN': '🟢',

        # Warning emojis
        'WARNING': '🟡',
        'AT RISK': '🟠',
        'NEEDS IMPROVEMENT': '🟠',
        'YELLOW': '🟡',

        # Critical emojis
        'CRITICAL': '🔴',
        'FAIL': '🔴',
        'POOR': '🔴',
        'OFF TRACK': '🔴',
        'RED': '🔴',
    }

    return emoji_map.get(status_upper, '⚪')


# ============================================================================
# CELL STYLING FUNCTIONS
# ============================================================================

def apply_header_style(
    cell,
    bold: bool = True,
    bg_color: str = COLOR_HEADER,
    text_color: str = COLOR_WHITE,
    font_size: int = 11,
    alignment: str = 'center'
):
    """
    Apply professional header styling to a cell.

    Args:
        cell: openpyxl cell object
        bold: Whether to make text bold
        bg_color: Background color (hex code)
        text_color: Text color (hex code)
        font_size: Font size in points
        alignment: Text alignment ('left', 'center', 'right')
    """
    cell.font = Font(bold=bold, color=text_color, size=font_size)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal=alignment, vertical='center', wrap_text=True)

    # Add borders
    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )
    cell.border = thin_border


def apply_cell_style(
    cell,
    bg_color: Optional[str] = None,
    bold: bool = False,
    align: str = 'left',
    number_format: Optional[str] = None,
    font_size: int = 10,
    text_color: Optional[str] = None
):
    """
    Apply general cell styling with optional formatting.

    Args:
        cell: openpyxl cell object
        bg_color: Background color (hex code, optional)
        bold: Whether to make text bold
        align: Text alignment ('left', 'center', 'right')
        number_format: Number format string (e.g., '0.0%', '#,##0')
        font_size: Font size in points
        text_color: Text color (hex code, optional)
    """
    # Apply font styling
    font_kwargs = {'size': font_size}
    if bold:
        font_kwargs['bold'] = True
    if text_color:
        font_kwargs['color'] = text_color
    cell.font = Font(**font_kwargs)

    # Apply background fill
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

    # Apply alignment
    cell.alignment = Alignment(horizontal=align, vertical='center')

    # Apply number formatting
    if number_format:
        cell.number_format = number_format

    # Add borders
    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )
    cell.border = thin_border


def apply_title_style(
    cell,
    font_size: int = 20,
    bold: bool = True,
    bg_color: Optional[str] = COLOR_HEADER,
    text_color: str = COLOR_WHITE
):
    """
    Apply title/heading style to a cell for report headers.

    Args:
        cell: openpyxl cell object
        font_size: Font size in points
        bold: Whether to make text bold
        bg_color: Background color (hex code, optional)
        text_color: Text color (hex code)
    """
    cell.font = Font(size=font_size, bold=bold, color=text_color)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal='left', vertical='center')


def apply_border(cell, style: str = 'thin', color: str = COLOR_BORDER):
    """
    Apply borders to a cell.

    Args:
        cell: openpyxl cell object
        style: Border style ('thin', 'medium', 'thick')
        color: Border color (hex code)
    """
    border = Border(
        left=Side(style=style, color=color),
        right=Side(style=style, color=color),
        top=Side(style=style, color=color),
        bottom=Side(style=style, color=color)
    )
    cell.border = border


# ============================================================================
# WORKSHEET UTILITY FUNCTIONS
# ============================================================================

def auto_adjust_column_width(
    worksheet,
    min_width: int = 12,
    max_width: int = 50
):
    """
    Auto-adjust column widths based on content.

    Iterates through all columns and sets width based on the longest
    content in each column, with min/max constraints.

    Args:
        worksheet: openpyxl worksheet object
        min_width: Minimum column width in characters
        max_width: Maximum column width in characters
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_status_cell(
    cell,
    status: str,
    include_emoji: bool = True,
    bold: bool = True,
    align: str = 'center'
):
    """
    Create a formatted status cell with color and optional emoji.

    Convenience function that combines status mapping, emoji, and styling.

    Args:
        cell: openpyxl cell object
        status: Status string (e.g., 'Met', 'Critical', 'GREEN')
        include_emoji: Whether to prepend status emoji
        bold: Whether to make text bold
        align: Text alignment

    Example:
        >>> create_status_cell(ws['A1'], 'Met')
        # Cell now contains "🟢 MET" with green background
    """
    # Get color and emoji for status
    bg_color = get_status_color(status)
    emoji = get_status_emoji(status) if include_emoji else ''

    # Set cell value with emoji
    display_text = f"{emoji} {status.upper()}" if emoji else status.upper()
    cell.value = display_text

    # Apply styling
    apply_cell_style(
        cell,
        bg_color=bg_color,
        bold=bold,
        align=align
    )


def freeze_header_rows(worksheet, row_count: int = 1):
    """
    Freeze top rows in a worksheet for scrolling.

    Args:
        worksheet: openpyxl worksheet object
        row_count: Number of rows to freeze (default: 1)
    """
    freeze_cell = f"A{row_count + 1}"
    worksheet.freeze_panes = freeze_cell


# ============================================================================
# LEGACY COMPATIBILITY
# ============================================================================
# These color constants are kept for backwards compatibility with existing code

COLORS = {
    'PASS': COLOR_LIGHT_GREEN,
    'FAIL': COLOR_LIGHT_RED,
    'EXCELLENT': COLOR_LIGHT_GREEN,
    'GOOD': COLOR_PALE_GREEN,
    'NEEDS IMPROVEMENT': COLOR_ORANGE,
    'POOR': COLOR_LIGHT_RED,
    'WARNING': COLOR_ORANGE,
    'CRITICAL': COLOR_RED,
    'HEADER': COLOR_HEADER,
    'SUBHEADER': COLOR_SUBHEADER,
}
