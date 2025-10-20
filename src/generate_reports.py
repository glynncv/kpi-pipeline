"""
generate_reports.py - Excel Report Generation Module

This module generates professional Excel reports from KPI calculation results.
Creates multi-sheet workbooks with formatting, charts, and visualizations.

Author: KPI Pipeline Project
Date: 2025-10-20
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference


class ReportGenerator:
    """
    Generate Excel reports from KPI results with professional formatting.
    
    Creates multi-sheet workbooks with:
    - Executive summary dashboard
    - KPI scorecard with status indicators
    - Individual KPI detail sheets
    - Raw data sheets for incidents and requests
    - Charts and visualizations
    - Color-coded status indicators
    
    Attributes:
        config: Configuration object with KPI targets and thresholds
    """
    
    # Color scheme for status indicators
    COLORS = {
        'PASS': '90EE90',      # Light green
        'FAIL': 'FFB6C1',      # Light red
        'EXCELLENT': '90EE90',  # Light green
        'GOOD': 'B4F8C8',      # Pale green
        'NEEDS IMPROVEMENT': 'FFD580',  # Light orange
        'POOR': 'FFB6C1',      # Light red
        'WARNING': 'FFD580',   # Light orange
        'CRITICAL': 'FFB6C1',  # Light red
        'HEADER': '4472C4',    # Blue
        'SUBHEADER': '70AD47', # Green
    }
    
    def __init__(self, config):
        """
        Initialize report generator with configuration.
        
        Args:
            config: Config object with KPI definitions and targets
        """
        self.config = config
    
    def generate_excel_report(self,
                            kpi_results: Dict[str, Any],
                            okr_results: Dict[str, Any],
                            action_triggers: Dict[str, list],
                            incidents: pd.DataFrame,
                            requests: pd.DataFrame,
                            output_path: str) -> None:
        """
        Generate complete Excel dashboard with all KPI and OKR results.
        
        Creates a multi-sheet workbook with:
        - Executive Summary (KPI + OKR)
        - KPI Scorecard
        - Individual KPI detail sheets
        - OKR Summary
        - Key Results Detail
        - Action Items
        
        Args:
            kpi_results: Dictionary of KPI results from calculate_kpis
            okr_results: Dictionary of OKR results from OKRCalculator
            action_triggers: Dictionary of action triggers (critical/warning)
            incidents: Transformed incidents DataFrame
            requests: Transformed requests DataFrame
            output_path: Path where Excel file should be saved
            
        Raises:
            Exception: If report generation fails
        """
        try:
            # Create output directory if needed
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Create workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create sheets in order
            self._create_executive_summary_sheet(wb, kpi_results, okr_results)
            self._create_scorecard_sheet(wb, kpi_results)
            self._create_kpi_detail_sheets(wb, kpi_results, incidents, requests)
            self._create_okr_summary_sheet(wb, okr_results)
            self._create_key_results_detail_sheet(wb, okr_results)
            self._create_action_items_sheet(wb, action_triggers, okr_results)
            # Note: Raw data sheets omitted for executive reporting
            # Operational analysis module will include detailed data sheets
            
            # Save workbook
            wb.save(output_path)
            print(f"âœ… Report saved: {output_path}")
            
        except Exception as e:
            raise Exception(f"Failed to generate Excel report: {str(e)}")
    
    def _create_executive_summary_sheet(self, 
                                       workbook: openpyxl.Workbook,
                                       kpi_results: Dict[str, Any],
                                       okr_results: Dict[str, Any]) -> None:
        """
        Create executive summary sheet with overall KPI and OKR scorecard.
        
        Shows:
        - Overall KPI score (large, prominent)
        - Overall OKR score (large, prominent)
        - Performance band with color coding
        - Key metrics dashboard
        - Weighted breakdown
        
        Args:
            workbook: openpyxl Workbook object
            kpi_results: Dictionary with all KPI results including OVERALL
            okr_results: Dictionary with OKR R002 results
        """
        ws = workbook.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = 'KPI EXECUTIVE SUMMARY'
        ws['A1'].font = Font(size=20, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['HEADER'],
                                     end_color=self.COLORS['HEADER'],
                                     fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30
        
        # Timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:E2')
        
        # Overall Score (large display)
        overall = kpi_results.get('OVERALL', {})
        score = overall.get('Overall_Score', 0)
        status = overall.get('Overall_Status', 'UNKNOWN')
        
        ws['A4'] = 'Overall Score'
        ws['A4'].font = Font(size=14, bold=True)
        
        ws['A5'] = f"{score:.1f}%"
        ws['A5'].font = Font(size=48, bold=True, color='FFFFFF')
        ws['A5'].fill = PatternFill(start_color=self.COLORS.get(status, 'CCCCCC'),
                                     end_color=self.COLORS.get(status, 'CCCCCC'),
                                     fill_type='solid')
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[5].height = 60
        
        ws['A6'] = status
        ws['A6'].font = Font(size=16, bold=True)
        ws['A6'].alignment = Alignment(horizontal='center')
        
        # Key Metrics Table
        ws['C4'] = 'Key Metrics'
        ws['C4'].font = Font(size=14, bold=True)
        
        metrics_data = [
            ['KPI', 'Status', 'Score'],
        ]
        
        # Add each KPI
        for kpi_code in ['SM001', 'SM002/KR4', 'SM003/KR5', 'SM004/KR6']:
            if kpi_code in kpi_results:
                kpi = kpi_results[kpi_code]
                metrics_data.append([
                    kpi.get('KPI_Name', kpi_code),
                    kpi.get('Status', ''),
                    f"{kpi.get('Adherence_Rate', 0):.1f}%"
                ])
        
        # Write metrics table
        row = 5
        for data_row in metrics_data:
            for col, value in enumerate(data_row, start=3):  # Start at column C
                cell = ws.cell(row=row, column=col, value=value)
                if row == 5:  # Header row
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                          end_color=self.COLORS['SUBHEADER'],
                                          fill_type='solid')
                elif col == 4:  # Status column
                    status = data_row[1]
                    if status in self.COLORS:
                        cell.fill = PatternFill(start_color=self.COLORS[status],
                                              end_color=self.COLORS[status],
                                              fill_type='solid')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center')
            row += 1
        
        # Weighted Breakdown
        ws['A9'] = 'Weighted Breakdown'
        ws['A9'].font = Font(size=14, bold=True)
        
        breakdown_data = [
            ['KPI', 'Score', 'Weight', 'Contribution'],
        ]
        
        weights_used = overall.get('Weights_Used', {})
        
        for kpi_code, weight in weights_used.items():
            if kpi_code in kpi_results:
                kpi = kpi_results[kpi_code]
                score = kpi.get('Adherence_Rate', 0)
                contribution = score * weight / 100
                breakdown_data.append([
                    kpi_code,
                    f"{score:.1f}%",
                    f"{weight}%",
                    f"{contribution:.1f}"
                ])
        
        # Write breakdown table
        row = 10
        for data_row in breakdown_data:
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                if row == 10:  # Header row
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                          end_color=self.COLORS['SUBHEADER'],
                                          fill_type='solid')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center')
            row += 1
        
        # Add OKR Score display (right side of executive summary)
        ws['G4'] = 'OKR R002 Score'
        ws['G4'].font = Font(size=14, bold=True)
        
        okr_score = okr_results.get('overall_score', 0)
        okr_status = okr_results.get('overall_status', 'UNKNOWN')
        
        ws['G5'] = f"{okr_score:.1f}%"
        ws['G5'].font = Font(size=48, bold=True, color='FFFFFF')
        # Strip emoji from status for color lookup
        status_clean = okr_status.split()[-1] if ' ' in okr_status else okr_status
        ws['G5'].fill = PatternFill(start_color=self.COLORS.get(status_clean, 'CCCCCC'),
                                     end_color=self.COLORS.get(status_clean, 'CCCCCC'),
                                     fill_type='solid')
        ws['G5'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['G6'] = okr_status
        ws['G6'].font = Font(size=14, bold=True)
        ws['G6'].alignment = Alignment(horizontal='center')
        
        # Auto-fit columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 20
    
    def _create_scorecard_sheet(self,
                               workbook: openpyxl.Workbook,
                               kpi_results: Dict[str, Any]) -> None:
        """
        Create KPI scorecard sheet with all KPIs in table format.
        
        Args:
            workbook: openpyxl Workbook object
            kpi_results: Dictionary with all KPI results
        """
        ws = workbook.create_sheet("KPI Scorecard")
        
        # Title
        ws['A1'] = 'KPI SCORECARD'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['HEADER'],
                                     end_color=self.COLORS['HEADER'],
                                     fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        # Headers
        headers = ['KPI Code', 'KPI Name', 'Status', 'Score', 'Target', 'Notes']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                   end_color=self.COLORS['SUBHEADER'],
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Data rows
        row = 4
        for kpi_code in ['SM001', 'SM002/KR4', 'SM003/KR5', 'SM004/KR6']:
            if kpi_code not in kpi_results:
                continue
                
            kpi = kpi_results[kpi_code]
            
            # Determine target and notes based on KPI type
            if kpi_code == 'SM001':
                target = f"P1â‰¤{kpi.get('P1_Target', 0)}, P2â‰¤{kpi.get('P2_Target', 5)}"
                notes = f"P1={kpi.get('P1_Count', 0)}, P2={kpi.get('P2_Count', 0)}"
            elif kpi_code in ['SM002/KR4', 'SM003/KR5']:
                target = f"â‰¥{kpi.get('Target_Adherence', 90)}%"
                if 'Backlog_Count' in kpi:
                    notes = f"{kpi.get('Backlog_Count', 0)} backlog / {kpi.get('Total_Incidents', 0)} total"
                else:
                    notes = f"{kpi.get('Aged_Count', 0)} aged / {kpi.get('Total_Requests', 0)} total"
            elif kpi_code == 'SM004/KR6':
                target = f"â‰¥{kpi.get('Target_Rate', 80)}%"
                notes = f"{kpi.get('FCR_Count', 0)} FCR / {kpi.get('Total_Resolved', 0)} resolved"
            else:
                target = ""
                notes = ""
            
            data = [
                kpi_code,
                kpi.get('KPI_Name', ''),
                kpi.get('Status', ''),
                f"{kpi.get('Adherence_Rate', 0):.1f}%",
                target,
                notes
            ]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                
                # Color code status column
                if col == 3:  # Status column
                    status = data[2]
                    if status in self.COLORS:
                        cell.fill = PatternFill(start_color=self.COLORS[status],
                                              end_color=self.COLORS[status],
                                              fill_type='solid')
                
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='left' if col in [2, 6] else 'center')
            
            row += 1
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40
    
    def _create_kpi_detail_sheets(self,
                                 workbook: openpyxl.Workbook,
                                 kpi_results: Dict[str, Any],
                                 incidents: pd.DataFrame,
                                 requests: pd.DataFrame) -> None:
        """
        Create individual detail sheets for each KPI.
        
        Args:
            workbook: openpyxl Workbook object
            kpi_results: Dictionary with all KPI results
            incidents: Incidents DataFrame
            requests: Requests DataFrame
        """
        # SM001 - Major Incidents
        if 'SM001' in kpi_results:
            self._create_sm001_sheet(workbook, kpi_results['SM001'], incidents)
        
        # SM002 - ServiceNow Backlog
        if 'SM002/KR4' in kpi_results:
            self._create_sm002_sheet(workbook, kpi_results['SM002/KR4'], incidents)
        
        # SM003/KR5 - Request Aging
        if 'SM003/KR5' in kpi_results:
            self._create_kr5_sheet(workbook, kpi_results['SM003/KR5'], requests)
        
        # SM004 - First Time Fix
        if 'SM004/KR6' in kpi_results:
            self._create_sm004_sheet(workbook, kpi_results['SM004/KR6'], incidents)
    
    def _create_sm001_sheet(self,
                           workbook: openpyxl.Workbook,
                           sm001: Dict[str, Any],
                           incidents: pd.DataFrame) -> None:
        """Create SM001 detail sheet"""
        ws = workbook.create_sheet("SM001 - Major Incidents")
        
        # Title
        ws['A1'] = 'SM001 - MAJOR INCIDENTS (P1 & P2)'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['HEADER'],
                                     end_color=self.COLORS['HEADER'],
                                     fill_type='solid')
        ws.merge_cells('A1:D1')
        
        # Summary metrics
        metrics = [
            ['Metric', 'Count', 'Target', 'Status'],
            ['P1 Incidents', sm001.get('P1_Count', 0), f"â‰¤{sm001.get('P1_Target', 0)}", 
             'PASS' if sm001.get('P1_Count', 0) <= sm001.get('P1_Target', 0) else 'FAIL'],
            ['P2 Incidents', sm001.get('P2_Count', 0), f"â‰¤{sm001.get('P2_Target', 5)}", 
             'PASS' if sm001.get('P2_Count', 0) <= sm001.get('P2_Target', 5) else 'FAIL'],
            ['Total Major', sm001.get('Total_Major', 0), f"â‰¤{sm001.get('P2_Target', 5)}", 
             sm001.get('Status', '')],
        ]
        
        for row_idx, row_data in enumerate(metrics, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                          end_color=self.COLORS['SUBHEADER'],
                                          fill_type='solid')
                elif col_idx == 4:  # Status column
                    if value in self.COLORS:
                        cell.fill = PatternFill(start_color=self.COLORS[value],
                                              end_color=self.COLORS[value],
                                              fill_type='solid')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center')
        
        # Column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
    
    def _create_sm002_sheet(self,
                           workbook: openpyxl.Workbook,
                           sm002: Dict[str, Any],
                           incidents: pd.DataFrame) -> None:
        """Create SM002 backlog detail sheet"""
        ws = workbook.create_sheet("SM002 - Backlog Analysis")
        
        # Title
        ws['A1'] = 'SM002 - SERVICENOW BACKLOG (KR4)'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['HEADER'],
                                     end_color=self.COLORS['HEADER'],
                                     fill_type='solid')
        ws.merge_cells('A1:D1')
        
        # Summary metrics
        total = sm002.get('Total_Incidents', 0)
        backlog = sm002.get('Backlog_Count', 0)
        non_backlog = total - backlog
        adherence = sm002.get('Backlog_Percentage', 0)
        target = sm002.get('Target_Adherence', 90)
        
        metrics = [
            ['Metric', 'Count', 'Percentage', 'Status'],
            ['Total Incidents', total, '100.0%', ''],
            ['Non-Backlog (<10 days)', non_backlog, f'{100-adherence:.1f}%', 
             'PASS' if adherence <= (100-target) else 'FAIL'],
            ['Backlog (â‰¥10 days)', backlog, f'{adherence:.1f}%', 
             'FAIL' if backlog > 0 else 'PASS'],
            ['', '', '', ''],
            ['Adherence Rate', '', f'{100-adherence:.1f}%', sm002.get('Status', '')],
            ['Target', '', f'â‰¥{target}%', ''],
        ]
        
        for row_idx, row_data in enumerate(metrics, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                          end_color=self.COLORS['SUBHEADER'],
                                          fill_type='solid')
                elif row_idx in [8, 9]:  # Summary rows
                    cell.font = Font(bold=True)
                elif col_idx == 4 and value in self.COLORS:  # Status
                    cell.fill = PatternFill(start_color=self.COLORS[value],
                                          end_color=self.COLORS[value],
                                          fill_type='solid')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center')
        
        # Column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 25
    
    def _create_kr5_sheet(self,
                         workbook: openpyxl.Workbook,
                         kr5: Dict[str, Any],
                         requests: pd.DataFrame) -> None:
        """Create KR5 request aging detail sheet"""
        ws = workbook.create_sheet("SM004 - Request Aging")
        
        # Title
        ws['A1'] = 'SM003 - SERVICE REQUEST AGING (KR5)'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['HEADER'],
                                     end_color=self.COLORS['HEADER'],
                                     fill_type='solid')
        ws.merge_cells('A1:D1')
        
        # Summary metrics
        total = kr5.get('Total_Requests', 0)
        aged = kr5.get('Aged_Count', 0)
        non_aged = total - aged
        adherence = kr5.get('Aged_Percentage', 0)
        target = kr5.get('Target_Adherence', 90)
        
        metrics = [
            ['Metric', 'Count', 'Percentage', 'Status'],
            ['Total Requests', total, '100.0%', ''],
            ['Non-Aged (<30 days)', non_aged, f'{100-adherence:.1f}%', 
             'PASS' if adherence <= (100-target) else 'FAIL'],
            ['Aged (â‰¥30 days)', aged, f'{adherence:.1f}%', 
             'FAIL' if aged > 0 else 'PASS'],
            ['', '', '', ''],
            ['Adherence Rate', '', f'{100-adherence:.1f}%', kr5.get('Status', '')],
            ['Target', '', f'â‰¥{target}%', ''],
        ]
        
        for row_idx, row_data in enumerate(metrics, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                          end_color=self.COLORS['SUBHEADER'],
                                          fill_type='solid')
                elif row_idx in [8, 9]:  # Summary rows
                    cell.font = Font(bold=True)
                elif col_idx == 4 and value in self.COLORS:  # Status
                    cell.fill = PatternFill(start_color=self.COLORS[value],
                                          end_color=self.COLORS[value],
                                          fill_type='solid')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center')
        
        # Column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 25
    
    def _create_sm004_sheet(self,
                           workbook: openpyxl.Workbook,
                           sm004: Dict[str, Any],
                           incidents: pd.DataFrame) -> None:
        """Create SM004 first time fix detail sheet"""
        ws = workbook.create_sheet("SM004 - First Time Fix")
        
        # Title
        ws['A1'] = 'SM004 - FIRST TIME FIX RATE (KR6)'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['HEADER'],
                                     end_color=self.COLORS['HEADER'],
                                     fill_type='solid')
        ws.merge_cells('A1:D1')
        
        # Summary metrics
        total_resolved = sm004.get('Total_Resolved', 0)
        fcr_count = sm004.get('FCR_Count', 0)
        fcr_rate = sm004.get('FCR_Percentage', 0)
        target = sm004.get('Target_Rate', 80)
        
        metrics = [
            ['Metric', 'Count', 'Percentage', 'Status'],
            ['Total Resolved', total_resolved, '100.0%', ''],
            ['First Call Resolution', fcr_count, f'{fcr_rate:.1f}%', 
             'PASS' if fcr_rate >= target else 'FAIL'],
            ['Reassigned', total_resolved - fcr_count, f'{100-fcr_rate:.1f}%', 
             'FAIL' if fcr_rate < target else ''],
            ['', '', '', ''],
            ['FCR Rate', '', f'{fcr_rate:.1f}%', sm004.get('Status', '')],
            ['Target', '', f'â‰¥{target}%', ''],
        ]
        
        for row_idx, row_data in enumerate(metrics, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                          end_color=self.COLORS['SUBHEADER'],
                                          fill_type='solid')
                elif row_idx in [8, 9]:  # Summary rows
                    cell.font = Font(bold=True)
                elif col_idx == 4 and value in self.COLORS:  # Status
                    cell.fill = PatternFill(start_color=self.COLORS[value],
                                          end_color=self.COLORS[value],
                                          fill_type='solid')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center')
        
        # Column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 25
    
    def _create_data_sheets(self,
                          workbook: openpyxl.Workbook,
                          incidents: pd.DataFrame,
                          requests: pd.DataFrame) -> None:
        """
        Create raw data sheets for incidents and requests.
        
        Args:
            workbook: openpyxl Workbook object
            incidents: Incidents DataFrame with calculated fields
            requests: Requests DataFrame with calculated fields
        """
        # Incident Details sheet
        ws_incidents = workbook.create_sheet("Incident Details")
        
        # Select key columns for incident details
        incident_cols = ['number', 'priority', 'incident_state', 'opened_at', 
                        'u_resolved', 'reassignment_count', 'location',
                        'Priority_Number', 'Is_Major_Incident', 'Is_Backlog', 'Is_First_Call_Resolution']
        
        available_incident_cols = [col for col in incident_cols if col in incidents.columns]
        incident_data = incidents[available_incident_cols].copy()
        
        # Write incident data
        for r_idx, row in enumerate(dataframe_to_rows(incident_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_incidents.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header row
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                          end_color=self.COLORS['SUBHEADER'],
                                          fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-fit columns
        for col in ws_incidents.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_incidents.column_dimensions[column].width = adjusted_width
        
        # Freeze header row
        ws_incidents.freeze_panes = 'A2'
        
        # Request Details sheet (only if requests data exists)
        if not requests.empty:
            ws_requests = workbook.create_sheet("Request Details")
            
            # Select key columns for request details
            request_cols = ['number', 'opened_at', 'closed_at', 'state',
                           'assignment_group', 'Is_Aged']
            
            available_request_cols = [col for col in request_cols if col in requests.columns]
            request_data = requests[available_request_cols].copy()
            
            # Write request data
            for r_idx, row in enumerate(dataframe_to_rows(request_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_requests.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:  # Header row
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                              end_color=self.COLORS['SUBHEADER'],
                                              fill_type='solid')
                        cell.alignment = Alignment(horizontal='center')
            
            # Auto-fit columns
            for col in ws_requests.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_requests.column_dimensions[column].width = adjusted_width
            
            # Freeze header row
            ws_requests.freeze_panes = 'A2'


    def _create_okr_summary_sheet(self,
                                  workbook: openpyxl.Workbook,
                                  okr_results: Dict[str, Any]) -> None:
        """Create OKR R002 summary sheet"""
        ws = workbook.create_sheet("OKR R002 Summary")
        
        # Title
        ws['A1'] = 'OKR R002 - SERVICE DELIVERY EXCELLENCE'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['HEADER'],
                                     end_color=self.COLORS['HEADER'],
                                     fill_type='solid')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 25
        
        # Objective
        ws['A3'] = 'Objective:'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = okr_results.get('objective', '')
        ws.merge_cells('B3:E3')
        
        # Overall Score
        ws['A5'] = 'Overall OKR Score'
        ws['A5'].font = Font(size=14, bold=True)
        ws['B5'] = f"{okr_results.get('overall_score', 0):.1f}%"
        ws['B5'].font = Font(size=18, bold=True)
        
        ws['A6'] = 'Status'
        ws['A6'].font = Font(bold=True)
        ws['B6'] = okr_results.get('overall_status', '')
        ws['B6'].font = Font(size=14, bold=True)
        
        # Key Results Summary Table
        ws['A9'] = 'Key Results Summary'
        ws['A9'].font = Font(size=14, bold=True)
        
        headers = ['KR ID', 'Name', 'Score', 'Status', 'Current', 'Target', 'Gap']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=10, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                   end_color=self.COLORS['SUBHEADER'],
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Add Key Results data
        row = 11
        for kr_id in ['KR3', 'KR4', 'KR5', 'KR6']:
            kr = okr_results['key_results'][kr_id]
            data = [
                kr_id,
                kr['name'],
                f"{kr['score']:.1f}%",
                kr['status'],
                f"{kr['current_value']:.1f}",
                f"{kr['target_operator']} {kr['target_value']}",
                f"{kr['gap_to_target']:.1f}"
            ]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center' if col != 2 else 'left')
            row += 1
        
        # Auto-fit columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['B'].width = 35
    
    def _create_key_results_detail_sheet(self,
                                        workbook: openpyxl.Workbook,
                                        okr_results: Dict[str, Any]) -> None:
        """Create Key Results detail sheet"""
        ws = workbook.create_sheet("Key Results Detail")
        
        # Title
        ws['A1'] = 'KEY RESULTS DETAIL'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['HEADER'],
                                     end_color=self.COLORS['HEADER'],
                                     fill_type='solid')
        ws.merge_cells('A1:E1')
        
        row = 3
        for kr_id in ['KR3', 'KR4', 'KR5', 'KR6']:
            kr = okr_results['key_results'][kr_id]
            
            # KR Header
            ws[f'A{row}'] = f"{kr_id}: {kr['name']}"
            ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
            ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                            end_color=self.COLORS['SUBHEADER'],
                                            fill_type='solid')
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # KR Details
            details = [
                ['Description:', kr['description']],
                ['Current Value:', f"{kr['current_value']:.1f}"],
                ['Target:', f"{kr['target_operator']} {kr['target_value']}"],
                ['Score:', f"{kr['score']:.1f}% / 100%"],
                ['Status:', kr['status']],
                ['Gap to Target:', f"{kr['gap_to_target']:.1f}"],
                ['Owner:', kr['owner']],
                ['Deadline:', kr['deadline']],
                ['Days Remaining:', kr['days_remaining']],
                ['Business Impact:', kr['business_impact']],
            ]
            
            for label, value in details:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                ws.merge_cells(f'B{row}:E{row}')
                row += 1
            
            row += 1  # Blank row between KRs
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
    
    def _create_action_items_sheet(self,
                                   workbook: openpyxl.Workbook,
                                   action_triggers: Dict[str, list],
                                   okr_results: Dict[str, Any]) -> None:
        """Create Action Items sheet"""
        ws = workbook.create_sheet("Action Items")
        
        # Title
        ws['A1'] = 'ACTION ITEMS & ESCALATIONS'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['HEADER'],
                                     end_color=self.COLORS['HEADER'],
                                     fill_type='solid')
        ws.merge_cells('A1:E1')
        
        # Critical Actions
        row = 3
        if action_triggers['critical']:
            ws[f'A{row}'] = 'ðŸ”´ CRITICAL ACTIONS REQUIRED'
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS['CRITICAL'],
                                            end_color=self.COLORS['CRITICAL'],
                                            fill_type='solid')
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Headers
            headers = ['KR ID', 'Action Required', 'Escalate To', 'Priority']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                       end_color=self.COLORS['SUBHEADER'],
                                       fill_type='solid')
            row += 1
            
            # Critical action rows
            for trigger in action_triggers['critical']:
                data = [
                    trigger['kr_id'],
                    trigger['action'],
                    trigger['escalation'],
                    'IMMEDIATE'
                ]
                for col, value in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                row += 1
            
            row += 1  # Blank row
        
        # Warning Actions
        if action_triggers['warning']:
            ws[f'A{row}'] = 'ðŸŸ¡ WARNING ACTIONS'
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS['WARNING'],
                                            end_color=self.COLORS['WARNING'],
                                            fill_type='solid')
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Headers
            headers = ['KR ID', 'Action Required', 'Escalate To', 'Priority']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.COLORS['SUBHEADER'],
                                       end_color=self.COLORS['SUBHEADER'],
                                       fill_type='solid')
            row += 1
            
            # Warning action rows
            for trigger in action_triggers['warning']:
                data = [
                    trigger['kr_id'],
                    trigger['action'],
                    trigger['escalation'],
                    'MONITOR'
                ]
                for col, value in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                row += 1
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15


def generate_excel_report(kpi_results: Dict[str, Any],
                          okr_results: Dict[str, Any],
                          action_triggers: Dict[str, list],
                          incidents: pd.DataFrame,
                          requests: pd.DataFrame,
                          config,
                          output_path: str) -> None:
    """
    Convenience function to generate Excel report.
    
    Args:
        kpi_results: Dictionary of KPI results
        okr_results: Dictionary of OKR results
        action_triggers: Dictionary of action triggers
        incidents: Transformed incidents DataFrame
        requests: Transformed requests DataFrame
        config: Configuration object
        output_path: Path for output file
    """
    generator = ReportGenerator(config)
    generator.generate_excel_report(kpi_results, okr_results, action_triggers, 
                                   incidents, requests, output_path)