"""
Excel Export Module
===================
Exports KPI and OKR results to formatted Excel workbooks.

Sheets created:
1. Summary - Overall OKR and KPI scores
2. OKR Details - Individual Key Results
3. KPI Details - Individual KPI metrics
4. Action Items - Triggered actions and recommendations

Author: IT Service Management Team
Version: 1.0
"""

from datetime import datetime
from typing import Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


class ExcelExporter:
    """Export KPI and OKR results to Excel."""
    
    # Color scheme
    COLORS = {
        'excellent': 'C6EFCE',  # Light green
        'good': 'FFEB9C',       # Light yellow
        'at_risk': 'FFCCCC',    # Light orange
        'critical': 'FFC7CE',   # Light red
        'header': '4472C4',     # Blue
        'subheader': 'D9E1F2',  # Light blue
    }
    
    def __init__(self, kpi_results: Dict[str, Any], okr_result: Optional[Dict[str, Any]] = None):
        """
        Initialize Excel exporter.
        
        Args:
            kpi_results: Dictionary of KPI calculation results
            okr_result: Optional OKR calculation result
        """
        self.kpi_results = kpi_results
        self.okr_result = okr_result
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
    def export(self, filename: str) -> str:
        """
        Export results to Excel file.
        
        Args:
            filename: Output filename (without path)
            
        Returns:
            Full path to created file
        """
        # Generate timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Add timestamp to filename if not present
        if not any(char.isdigit() for char in filename):
            name_parts = filename.rsplit('.', 1)
            filename = f"{name_parts[0]}_{timestamp}.{name_parts[1]}" if len(name_parts) > 1 else f"{filename}_{timestamp}.xlsx"
        
        filepath = f"data/output/{filename}"
        
        # Create sheets
        self._create_summary_sheet()
        
        if self.okr_result:
            self._create_okr_details_sheet()
        
        self._create_kpi_details_sheet()
        
        if self.okr_result:
            self._create_action_items_sheet()
        
        # Save workbook
        self.wb.save(filepath)
        
        return filepath
    
    def _create_summary_sheet(self):
        """Create summary sheet with overall scores."""
        ws = self.wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = 'KPI & OKR Dashboard'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
        ws.merge_cells('A1:D1')
        
        # Timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:D2')
        
        row = 4
        
        # OKR Summary
        if self.okr_result:
            ws[f'A{row}'] = 'OKR R002: Service Delivery Excellence'
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            # Overall OKR Score
            ws[f'A{row}'] = 'Overall OKR Score:'
            ws[f'B{row}'] = self.okr_result['overall_score']
            ws[f'C{row}'] = '/100'
            ws[f'D{row}'] = self.okr_result['overall_status']
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(size=14, bold=True)
            ws[f'D{row}'].font = Font(bold=True)
            
            # Color code status
            status = self.okr_result['overall_status']
            fill_color = self._get_status_color(status)
            ws[f'D{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            row += 2
            
            # Key Results
            ws[f'A{row}'] = 'Key Result'
            ws[f'B{row}'] = 'Score'
            ws[f'C{row}'] = 'Status'
            ws[f'D{row}'] = 'Gap to Target'
            self._apply_header_style(ws, row, ['A', 'B', 'C', 'D'])
            row += 1
            
            for kr_id, kr_data in self.okr_result['key_results'].items():
                ws[f'A{row}'] = f"{kr_id}: {kr_data['name']}"
                ws[f'B{row}'] = kr_data['score']
                ws[f'C{row}'] = kr_data['status']
                ws[f'D{row}'] = kr_data['gap_to_target']
                
                fill_color = self._get_status_color(kr_data['status'])
                ws[f'C{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                row += 1
            
            row += 2
        
        # KPI Summary
        ws[f'A{row}'] = 'KPI Performance'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Overall KPI Score
        if 'OVERALL' in self.kpi_results:
            overall = self.kpi_results['OVERALL']
            ws[f'A{row}'] = 'Overall KPI Score:'
            ws[f'B{row}'] = overall['Overall_Score']
            ws[f'C{row}'] = '%'
            ws[f'D{row}'] = overall['Overall_Status']
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(size=14, bold=True)
            ws[f'D{row}'].font = Font(bold=True)
            row += 2
        
        # Individual KPIs
        ws[f'A{row}'] = 'KPI'
        ws[f'B{row}'] = 'Adherence'
        ws[f'C{row}'] = 'Status'
        ws[f'D{row}'] = 'Impact'
        self._apply_header_style(ws, row, ['A', 'B', 'C', 'D'])
        row += 1
        
        for kpi_code, kpi_data in self.kpi_results.items():
            if kpi_code == 'OVERALL':
                continue
            
            ws[f'A{row}'] = f"{kpi_code}: {kpi_data['KPI_Name']}"
            ws[f'B{row}'] = f"{kpi_data['Adherence_Rate']}%"
            ws[f'C{row}'] = kpi_data['Status']
            ws[f'D{row}'] = kpi_data['Business_Impact']
            row += 1
        
        # Auto-size columns
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 25
    
    def _create_okr_details_sheet(self):
        """Create detailed OKR sheet."""
        ws = self.wb.create_sheet("OKR Details")
        
        # Title
        ws['A1'] = 'OKR R002: Service Delivery Excellence'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
        ws.merge_cells('A1:F1')
        
        row = 3
        
        # Overall Summary
        ws[f'A{row}'] = 'Overall OKR Score'
        ws[f'B{row}'] = self.okr_result['overall_score']
        ws[f'C{row}'] = self.okr_result['overall_status']
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(size=12, bold=True)
        fill_color = self._get_status_color(self.okr_result['overall_status'])
        ws[f'C{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        row += 1
        
        ws[f'A{row}'] = 'Objective'
        ws[f'B{row}'] = self.okr_result['objective']
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'B{row}:F{row}')
        row += 2
        
        # Key Results Table
        ws[f'A{row}'] = 'Key Result'
        ws[f'B{row}'] = 'Current'
        ws[f'C{row}'] = 'Target'
        ws[f'D{row}'] = 'Score'
        ws[f'E{row}'] = 'Status'
        ws[f'F{row}'] = 'Deadline'
        self._apply_header_style(ws, row, ['A', 'B', 'C', 'D', 'E', 'F'])
        row += 1
        
        for kr_id, kr_data in self.okr_result['key_results'].items():
            ws[f'A{row}'] = f"{kr_id}: {kr_data['name']}"
            ws[f'B{row}'] = f"{kr_data['current_value']}"
            ws[f'C{row}'] = f"{kr_data['target_operator']}{kr_data['target_value']}"
            ws[f'D{row}'] = f"{kr_data['score']}/100"
            ws[f'E{row}'] = kr_data['status']
            ws[f'F{row}'] = f"{kr_data['deadline']} ({kr_data['days_remaining']}d)"
            
            fill_color = self._get_status_color(kr_data['status'])
            ws[f'E{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            row += 1
        
        row += 1
        
        # Additional Details
        ws[f'A{row}'] = 'Detailed Information'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        for kr_id, kr_data in self.okr_result['key_results'].items():
            ws[f'A{row}'] = kr_id
            ws[f'B{row}'] = kr_data['name']
            ws[f'A{row}'].font = Font(bold=True)
            ws.merge_cells(f'B{row}:F{row}')
            row += 1
            
            ws[f'B{row}'] = 'Description:'
            ws[f'C{row}'] = kr_data['description']
            ws.merge_cells(f'C{row}:F{row}')
            row += 1
            
            ws[f'B{row}'] = 'Owner:'
            ws[f'C{row}'] = kr_data['owner']
            row += 1
            
            ws[f'B{row}'] = 'Gap to Target:'
            ws[f'C{row}'] = kr_data['gap_to_target']
            row += 1
            
            ws[f'B{row}'] = 'Criticality:'
            ws[f'C{row}'] = kr_data['criticality']
            row += 2
        
        # Auto-size columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
    
    def _create_kpi_details_sheet(self):
        """Create detailed KPI sheet."""
        ws = self.wb.create_sheet("KPI Details")
        
        # Title
        ws['A1'] = 'KPI Detailed Metrics'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
        ws.merge_cells('A1:E1')
        
        row = 3
        
        for kpi_code, kpi_data in self.kpi_results.items():
            if kpi_code == 'OVERALL':
                continue
            
            # KPI Header
            ws[f'A{row}'] = f"{kpi_code}: {kpi_data['KPI_Name']}"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS['subheader'], end_color=self.COLORS['subheader'], fill_type='solid')
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # KPI Details
            ws[f'B{row}'] = 'Status:'
            ws[f'C{row}'] = kpi_data['Status']
            row += 1
            
            ws[f'B{row}'] = 'Adherence Rate:'
            ws[f'C{row}'] = f"{kpi_data['Adherence_Rate']}%"
            row += 1
            
            ws[f'B{row}'] = 'Business Impact:'
            ws[f'C{row}'] = kpi_data['Business_Impact']
            row += 1
            
            # KPI-specific metrics
            if 'P1_Count' in kpi_data:
                ws[f'B{row}'] = 'P1 Count:'
                ws[f'C{row}'] = f"{kpi_data['P1_Count']} (target: ≤{kpi_data['P1_Target']})"
                row += 1
                ws[f'B{row}'] = 'P2 Count:'
                ws[f'C{row}'] = f"{kpi_data['P2_Count']} (target: ≤{kpi_data['P2_Target']})"
                row += 1
                ws[f'B{row}'] = 'Total Major:'
                ws[f'C{row}'] = kpi_data['Total_Major']
                row += 1
            
            elif 'Backlog_Count' in kpi_data:
                ws[f'B{row}'] = 'Total Incidents:'
                ws[f'C{row}'] = kpi_data['Total_Incidents']
                row += 1
                ws[f'B{row}'] = 'Backlog Count:'
                ws[f'C{row}'] = kpi_data['Backlog_Count']
                row += 1
                ws[f'B{row}'] = 'Backlog Percentage:'
                ws[f'C{row}'] = f"{kpi_data['Backlog_Percentage']}%"
                row += 1
                ws[f'B{row}'] = 'Target Adherence:'
                ws[f'C{row}'] = f"≥{kpi_data['Target_Adherence']}%"
                row += 1
            
            elif 'Aged_Count' in kpi_data:
                ws[f'B{row}'] = 'Total Requests:'
                ws[f'C{row}'] = kpi_data['Total_Requests']
                row += 1
                ws[f'B{row}'] = 'Aged Count:'
                ws[f'C{row}'] = kpi_data['Aged_Count']
                row += 1
                ws[f'B{row}'] = 'Aged Percentage:'
                ws[f'C{row}'] = f"{kpi_data['Aged_Percentage']}%"
                row += 1
                ws[f'B{row}'] = 'Target Adherence:'
                ws[f'C{row}'] = f"≥{kpi_data['Target_Adherence']}%"
                row += 1
            
            elif 'FCR_Count' in kpi_data:
                ws[f'B{row}'] = 'Total Resolved:'
                ws[f'C{row}'] = kpi_data['Total_Resolved']
                row += 1
                ws[f'B{row}'] = 'FCR Count:'
                ws[f'C{row}'] = kpi_data['FCR_Count']
                row += 1
                ws[f'B{row}'] = 'FCR Percentage:'
                ws[f'C{row}'] = f"{kpi_data['FCR_Percentage']}%"
                row += 1
                ws[f'B{row}'] = 'Target Rate:'
                ws[f'C{row}'] = f"≥{kpi_data['Target_Rate']}%"
                row += 1
            
            row += 1  # Space between KPIs
        
        # Auto-size columns
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 25
    
    def _create_action_items_sheet(self):
        """Create action items sheet."""
        ws = self.wb.create_sheet("Action Items")
        
        # Title
        ws['A1'] = 'Recommended Actions'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Note: Action triggers would come from OKR calculator
        # For now, create based on KR status
        ws[f'A{row}'] = 'Priority'
        ws[f'B{row}'] = 'Key Result'
        ws[f'C{row}'] = 'Action Required'
        ws[f'D{row}'] = 'Owner'
        self._apply_header_style(ws, row, ['A', 'B', 'C', 'D'])
        row += 1
        
        # Generate actions based on KR scores
        for kr_id, kr_data in self.okr_result['key_results'].items():
            if kr_data['score'] < 50:
                priority = '🔴 CRITICAL'
                fill_color = self.COLORS['critical']
            elif kr_data['score'] < 70:
                priority = '🟠 HIGH'
                fill_color = self.COLORS['at_risk']
            elif kr_data['score'] < 90:
                priority = '🟡 MEDIUM'
                fill_color = self.COLORS['good']
            else:
                continue  # No action needed for excellent performance
            
            ws[f'A{row}'] = priority
            ws[f'B{row}'] = f"{kr_id}: {kr_data['name']}"
            
            # Generate action based on KR
            if 'Backlog' in kr_data['name']:
                action = f"Reduce backlog from {kr_data['current_value']}% to {kr_data['target_value']}% (gap: {kr_data['gap_to_target']}%)"
            elif 'Fix Rate' in kr_data['name']:
                action = f"Improve rate from {kr_data['current_value']}% to {kr_data['target_value']}% (gap: {abs(kr_data['gap_to_target'])}%)"
            else:
                action = f"Close gap of {abs(kr_data['gap_to_target'])} to meet target"
            
            ws[f'C{row}'] = action
            ws[f'D{row}'] = kr_data['owner']
            
            ws[f'A{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 25
    
    def _apply_header_style(self, ws, row: int, columns: list):
        """Apply header styling to cells."""
        for col in columns:
            cell = ws[f'{col}{row}']
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _get_status_color(self, status: str) -> str:
        """Get fill color based on status."""
        status_lower = status.lower()
        if 'excellent' in status_lower or '🟢' in status:
            return self.COLORS['excellent']
        elif 'good' in status_lower or 'on track' in status_lower or '🟡' in status:
            return self.COLORS['good']
        elif 'at risk' in status_lower or 'warning' in status_lower or '🟠' in status:
            return self.COLORS['at_risk']
        else:
            return self.COLORS['critical']


def export_to_excel(kpi_results: Dict[str, Any], okr_result: Optional[Dict[str, Any]] = None, 
                   filename: str = "kpi_report.xlsx") -> str:
    """
    Export KPI and OKR results to Excel.
    
    Args:
        kpi_results: Dictionary of KPI calculation results
        okr_result: Optional OKR calculation result
        filename: Output filename
        
    Returns:
        Path to created Excel file
    """
    exporter = ExcelExporter(kpi_results, okr_result)
    filepath = exporter.export(filename)
    return filepath


if __name__ == "__main__":
    # Test with sample data
    print("Excel export module loaded successfully")
    print("Use export_to_excel(kpi_results, okr_result) to create reports")

