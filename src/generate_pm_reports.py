"""
Problem Management KPI Dashboard Export Module

This module generates professional Excel dashboards for Problem Management KPIs.
Creates multi-sheet workbooks with KPI summaries and detailed problem breakdowns.

Author: KPI Pipeline Project - Session 4
Date: 2025-11-04
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict, Any
import os

# Import src modules
try:
    from . import config_loader
    from . import load_problem_data
    from . import transform_problems
    from . import calculate_pm_kpis
    from .report_utils import (
        get_status_color,
        get_status_emoji,
        apply_header_style,
        apply_cell_style,
        auto_adjust_column_width,
        COLOR_HEADER,
        COLOR_RED,
        COLOR_GREEN,
        COLOR_YELLOW
    )
except ImportError:
    try:
        from src import config_loader
        from src import load_problem_data
        from src import transform_problems
        from src import calculate_pm_kpis
        from src.report_utils import (
            get_status_color,
            get_status_emoji,
            apply_header_style,
            apply_cell_style,
            auto_adjust_column_width,
            COLOR_HEADER,
            COLOR_RED,
            COLOR_GREEN,
            COLOR_YELLOW
        )
    except ImportError:
        import config_loader
        import load_problem_data
        import transform_problems
        import calculate_pm_kpis
        from report_utils import (
            get_status_color,
            get_status_emoji,
            apply_header_style,
            apply_cell_style,
            auto_adjust_column_width,
            COLOR_HEADER,
            COLOR_RED,
            COLOR_GREEN,
            COLOR_YELLOW
        )


# ============================================================================
# SUMMARY SHEET CREATION
# ============================================================================

def create_summary_sheet(workbook: Workbook, kpi_results: Dict[str, Any]) -> None:
    """
    Create the KPI Summary sheet with formatted metrics and status indicators.
    
    This sheet provides a high-level overview of the RCA001 KPI including:
    - Completion rate and target
    - Status indicator with color coding
    - Gap analysis
    - Detailed counts
    - Calculation metadata
    
    Args:
        workbook: openpyxl Workbook object
        kpi_results: Dictionary with KPI calculation results from calculate_all_pm_kpis()
                    Expected keys: 'kpi_id', 'completion_rate', 'target', 'status',
                                   'gap', 'completed_ontime', 'total_requiring_rca',
                                   'total_problems', 'calculation_date'
    
    Example:
        >>> wb = Workbook()
        >>> kpi_data = calculate_all_pm_kpis(df, config)
        >>> create_summary_sheet(wb, kpi_data)
        >>> wb.save('dashboard.xlsx')
    """
    # Get or create summary sheet
    if 'Summary' in workbook.sheetnames:
        ws = workbook['Summary']
    else:
        ws = workbook.create_sheet('Summary', 0)
    
    # Extract RCA001 results
    rca_results = kpi_results.get('RCA001', {})
    
    # ========================================================================
    # HEADER SECTION
    # ========================================================================
    
    # Title
    ws['A1'] = 'Problem Management KPI Dashboard'
    ws['A1'].font = Font(bold=True, size=16, color=COLOR_HEADER)
    ws.merge_cells('A1:E1')
    
    # Subtitle with date
    calc_date = rca_results.get('calculation_date', datetime.now().strftime('%Y-%m-%d'))
    ws['A2'] = f'Report Date: {calc_date}'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')
    
    # Add spacing
    ws.row_dimensions[3].height = 5
    
    # ========================================================================
    # KPI HEADER ROW
    # ========================================================================
    
    headers = ['KPI', 'Description', 'Actual', 'Target', 'Status', 'Gap', 'Performance']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_header_style(cell)
    
    # ========================================================================
    # KPI DATA ROW
    # ========================================================================
    
    row = 5
    
    # KPI ID
    ws.cell(row=row, column=1).value = rca_results.get('kpi_id', 'RCA001')
    apply_cell_style(ws.cell(row=row, column=1), bold=True)
    
    # Description
    ws.cell(row=row, column=2).value = 'Root Cause Analysis Completion Rate'
    apply_cell_style(ws.cell(row=row, column=2))
    
    # Actual (completion rate)
    actual = rca_results.get('completion_rate', 0)
    ws.cell(row=row, column=3).value = actual / 100  # Convert to decimal for percentage format
    apply_cell_style(ws.cell(row=row, column=3), align='center', number_format='0.0%')
    
    # Target
    target = rca_results.get('target', 95)
    ws.cell(row=row, column=4).value = target / 100  # Convert to decimal for percentage format
    apply_cell_style(ws.cell(row=row, column=4), align='center', number_format='0.0%')
    
    # Status (with emoji and color)
    status = rca_results.get('status', 'UNKNOWN')
    status_emoji = get_status_emoji(status)
    status_color = get_status_color(status)
    ws.cell(row=row, column=5).value = f"{status_emoji} {status}"
    apply_cell_style(ws.cell(row=row, column=5), bg_color=status_color, 
                    bold=True, align='center')
    
    # Gap
    gap = rca_results.get('gap', 0)
    ws.cell(row=row, column=6).value = gap / 100  # Convert to decimal for percentage format
    apply_cell_style(ws.cell(row=row, column=6), align='center', number_format='0.0%')
    
    # Performance (completed / total)
    completed = rca_results.get('completed_ontime', 0)
    total_rca = rca_results.get('total_requiring_rca', 0)
    ws.cell(row=row, column=7).value = f"{completed} / {total_rca}"
    apply_cell_style(ws.cell(row=row, column=7), align='center')
    
    # ========================================================================
    # DETAILS SECTION
    # ========================================================================
    
    # Add spacing
    ws.row_dimensions[6].height = 5
    
    # Details header
    ws['A7'] = 'Detailed Breakdown'
    ws['A7'].font = Font(bold=True, size=12, color=COLOR_HEADER)
    ws.merge_cells('A7:D7')
    
    # Details data
    details = [
        ('Total Problems (P1/P2):', rca_results.get('total_problems', 0)),
        ('Requiring RCA:', rca_results.get('total_requiring_rca', 0)),
        ('Completed On-Time:', rca_results.get('completed_ontime', 0)),
        ('Late or Incomplete:', total_rca - completed if total_rca > completed else 0)
    ]
    
    start_row = 8
    for idx, (label, value) in enumerate(details):
        row = start_row + idx
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
    
    # ========================================================================
    # THRESHOLD SECTION
    # ========================================================================
    
    # Add spacing
    ws.row_dimensions[12].height = 5
    
    # Thresholds header
    ws['A13'] = 'Status Thresholds'
    ws['A13'].font = Font(bold=True, size=12, color=COLOR_HEADER)
    ws.merge_cells('A13:C13')
    
    # Threshold details (these should come from config in real implementation)
    thresholds = [
        ('GREEN:', '≥ 95%'),
        ('YELLOW:', '≥ 85% and < 95%'),
        ('RED:', '< 85%')
    ]
    
    start_row = 14
    for idx, (label, criteria) in enumerate(thresholds):
        row = start_row + idx
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = criteria
    
    # ========================================================================
    # FORMATTING ADJUSTMENTS
    # ========================================================================
    
    # Auto-adjust column widths
    auto_adjust_column_width(ws, min_width=15, max_width=40)
    
    # Freeze top rows
    ws.freeze_panes = 'A5'
    
    print(f"[OK] Summary sheet created successfully")


# ============================================================================
# DETAIL SHEET CREATION
# ============================================================================

def create_detail_sheet(workbook: Workbook, df: pd.DataFrame) -> None:
    """
    Create the Problem Details sheet with all RCA-requiring problems.
    
    This sheet provides a detailed view of all problems that require RCA,
    including their status, timeline, and completion information.
    
    Args:
        workbook: openpyxl Workbook object
        df: Transformed DataFrame from transform_all_problem_data()
            Must include columns: number, priority, state, opened_at,
            Requires_RCA, RCA_OnTime, rca_stage, Days_Open, Is_Major_Problem
    
    Example:
        >>> wb = Workbook()
        >>> transformed_df = transform_all_problem_data(problems, tasks)
        >>> create_detail_sheet(wb, transformed_df)
        >>> wb.save('dashboard.xlsx')
    """
    # Get or create detail sheet
    if 'RCA Details' in workbook.sheetnames:
        ws = workbook['RCA Details']
    else:
        ws = workbook.create_sheet('RCA Details')
    
    # Filter to only problems requiring RCA
    detail_df = df[df['Requires_RCA'] == True].copy()
    
    # Select and order columns for the report
    # Use actual column names from transform_problems.py
    columns_to_show = [
        'number',           # Problem number
        'priority',         # Priority (e.g., "2 - High")
        'state',            # Problem state
        'opened_at',        # Created date
        'Days_Open',        # Days open
        'Requires_RCA',     # Requires RCA flag
        'RCA_OnTime',       # RCA on-time flag
        'rca_stage'         # RCA stage (if available)
    ]
    
    # Ensure all columns exist (use only available columns)
    available_columns = [col for col in columns_to_show if col in detail_df.columns]
    detail_df = detail_df[available_columns].copy()
    
    # Rename columns for better readability in Excel
    column_names = {
        'number': 'Problem Number',
        'priority': 'Priority',
        'state': 'State',
        'opened_at': 'Created Date',
        'Days_Open': 'Days Open',
        'Requires_RCA': 'Requires RCA',
        'RCA_OnTime': 'RCA On-Time',
        'rca_stage': 'RCA Stage'
    }
    
    # Rename only columns that exist
    rename_map = {k: v for k, v in column_names.items() if k in detail_df.columns}
    detail_df = detail_df.rename(columns=rename_map)
    
    # ========================================================================
    # HEADER SECTION
    # ========================================================================
    
    # Title
    ws['A1'] = 'Problem Details - RCA Required'
    ws['A1'].font = Font(bold=True, size=14, color=COLOR_HEADER)
    ws.merge_cells(f'A1:{get_column_letter(len(detail_df.columns))}1')
    
    # Subtitle
    ws['A2'] = f'Total Problems Requiring RCA: {len(detail_df)}'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells(f'A2:{get_column_letter(len(detail_df.columns))}2')
    
    # Add spacing
    ws.row_dimensions[3].height = 5
    
    # ========================================================================
    # WRITE DATA TO SHEET
    # ========================================================================
    
    # Write headers (row 4)
    for col_idx, column_name in enumerate(detail_df.columns, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = column_name
        apply_header_style(cell)
    
    # Write data (starting from row 5)
    start_row = 5
    for row_idx, row_data in enumerate(dataframe_to_rows(detail_df, index=False, header=False), start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # Apply formatting based on column
            column_name = detail_df.columns[col_idx - 1]
            
            if column_name == 'Problem Number':
                apply_cell_style(cell, bold=True)
            elif column_name == 'Priority':
                # Color code priorities (extract number from "2 - High")
                if pd.notna(value):
                    priority_str = str(value)
                    if '1' in priority_str or 'Critical' in priority_str:
                        apply_cell_style(cell, bg_color=COLOR_RED, align='center')
                    elif '2' in priority_str or 'High' in priority_str:
                        apply_cell_style(cell, bg_color=COLOR_YELLOW, align='center')
                    else:
                        apply_cell_style(cell, align='center')
                else:
                    apply_cell_style(cell, align='center')
            elif column_name == 'Days Open':
                apply_cell_style(cell, align='center')
            elif column_name == 'RCA On-Time':
                # Color code RCA status
                if value == True or value == 'True' or (isinstance(value, str) and value.lower() == 'true'):
                    apply_cell_style(cell, bg_color=COLOR_GREEN, align='center')
                elif value == False or value == 'False' or (isinstance(value, str) and value.lower() == 'false'):
                    apply_cell_style(cell, bg_color=COLOR_RED, align='center')
                else:
                    apply_cell_style(cell, align='center')
            elif column_name == 'Created Date':
                # Format dates
                if pd.notna(value):
                    try:
                        if isinstance(value, pd.Timestamp):
                            cell.value = value.strftime('%Y-%m-%d')
                        elif isinstance(value, str):
                            cell.value = value
                    except:
                        pass
                apply_cell_style(cell)
            else:
                apply_cell_style(cell)
    
    # ========================================================================
    # FORMATTING ADJUSTMENTS
    # ========================================================================
    
    # Auto-adjust column widths
    auto_adjust_column_width(ws, min_width=12, max_width=30)
    
    # Freeze top rows and first column
    ws.freeze_panes = 'B5'
    
    print(f"[OK] Detail sheet created with {len(detail_df)} problems")


# ============================================================================
# MAIN EXPORT FUNCTION
# ============================================================================

def export_pm_dashboard(kpi_results: Dict[str, Any], 
                       transformed_df: pd.DataFrame,
                       output_dir: str = 'data/output',
                       filename: str = None) -> str:
    """
    Create and export a complete Problem Management KPI dashboard to Excel.
    
    This is the main function that orchestrates the creation of a multi-sheet
    Excel workbook with KPI summary and problem details.
    
    Args:
        kpi_results: Dictionary with KPI calculation results from calculate_all_pm_kpis()
        transformed_df: Transformed DataFrame from transform_all_problem_data()
        output_dir: Directory to save the Excel file (default: 'data/output')
        filename: Optional custom filename (default: auto-generated with timestamp)
        
    Returns:
        Full path to the created Excel file
        
    Raises:
        ValueError: If required data is missing or invalid
        IOError: If unable to write the file
        
    Example:
        >>> # After calculating KPIs and transforming data
        >>> kpis = calculate_all_pm_kpis(transformed_df, config)
        >>> filepath = export_pm_dashboard(kpis, transformed_df)
        >>> print(f"Dashboard created: {filepath}")
        Dashboard created: data/output/PM_Dashboard_2025-11-04.xlsx
    """
    # ========================================================================
    # VALIDATION
    # ========================================================================
    
    if not kpi_results:
        raise ValueError("KPI results cannot be empty")
    
    if transformed_df is None or transformed_df.empty:
        raise ValueError("Transformed DataFrame cannot be empty")
    
    # Validate required columns
    required_columns = ['Requires_RCA', 'RCA_OnTime']
    missing_columns = [col for col in required_columns if col not in transformed_df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")
    
    # ========================================================================
    # SETUP OUTPUT PATH
    # ========================================================================
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename if not provided
    if filename is None:
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        filename = f'PM_Dashboard_{timestamp}.xlsx'
    
    # Ensure .xlsx extension
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    output_path = os.path.join(output_dir, filename)
    
    # ========================================================================
    # CREATE WORKBOOK
    # ========================================================================
    
    print(f"\n[INFO] Creating Problem Management Dashboard...")
    print(f"       Output: {output_path}")
    
    # Create new workbook
    wb = Workbook()
    
    # Remove default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========================================================================
    # CREATE SHEETS
    # ========================================================================
    
    print(f"\n[INFO] Creating Summary sheet...")
    create_summary_sheet(wb, kpi_results)
    
    print(f"[INFO] Creating Detail sheet...")
    create_detail_sheet(wb, transformed_df)
    
    # ========================================================================
    # SAVE WORKBOOK
    # ========================================================================
    
    try:
        wb.save(output_path)
        print(f"\n[OK] Dashboard created successfully!")
        print(f"     Location: {output_path}")
        print(f"     Sheets: {', '.join(wb.sheetnames)}")
        return output_path
    
    except PermissionError as e:
        error_msg = (
            f"Permission denied: Cannot save file '{output_path}'\n"
            f"  - File may be open in Excel or another program\n"
            f"  - Please close the file and try again\n"
            f"  - Or specify a different filename using the 'filename' parameter"
        )
        raise IOError(error_msg) from e
    
    except Exception as e:
        raise IOError(f"Failed to save Excel file: {str(e)}")


# ============================================================================
# TEST FUNCTION
# ============================================================================

def test_export():
    """
    Test function to validate the export module with sample data.
    
    This function creates a sample dashboard using the EMEA Problem Management data.
    It requires that all previous modules (Sessions 1-3) are available and working.
    
    Expected output:
    - Excel file created in output/ directory
    - Summary sheet with RCA001 KPI (74.4%, RED status)
    - Detail sheet with 39 problems requiring RCA
    
    To run this test:
        python src/generate_pm_reports.py
        
    Or from Python:
        from src.generate_pm_reports import test_export
        test_export()
    """
    print("=" * 80)
    print("TESTING: Problem Management Report Export")
    print("=" * 80)
    
    try:
        # Import required modules
        print("\n1. Importing modules...")
        
        # Load configuration
        print("\n2. Loading configuration...")
        config = config_loader.load_config()
        print(f"   [OK] Config loaded")
        
        # Load data
        print("\n3. Loading problem data...")
        problems_df, tasks_df = load_problem_data.load_all_problem_data('data/input', config)
        
        if problems_df is None or tasks_df is None:
            print("   [SKIP] Data files not available - creating sample dashboard")
            
            # Create sample data for testing
            current_date = pd.Timestamp('2025-11-03')
            sample_problems = pd.DataFrame({
                'number': ['PRB001', 'PRB002', 'PRB003'],
                'priority': ['2 - High', '2 - High', '1 - Critical'],
                'u_rca_required': ['Yes', 'Yes', 'Yes'],
                'opened_at': [current_date - pd.Timedelta(days=30)] * 3,
                'closed_at': [None] * 3,
                'state': ['Open', 'Pending Change', 'Open']
            })
            
            sample_tasks = pd.DataFrame({
                'task': ['PTASK001', 'PTASK002', 'PTASK003'],
                'task.parent.number': ['PRB001', 'PRB002', 'PRB003'],
                'stage': ['Achieved', 'Breached', 'In progress']
            })
            
            transformed_df = transform_problems.transform_all_problem_data(sample_problems, sample_tasks)
            kpi_results = calculate_pm_kpis.calculate_all_pm_kpis(transformed_df, config)
            
            print("   [OK] Using sample data for testing")
        else:
            print(f"   [OK] Loaded {len(problems_df)} problems and {len(tasks_df)} tasks")
            
            # Transform data
            print("\n4. Transforming data...")
            transformed_df = transform_problems.transform_all_problem_data(problems_df, tasks_df)
            print(f"   [OK] Transformed {len(transformed_df)} problems")
            
            # Calculate KPIs
            print("\n5. Calculating KPIs...")
            kpi_results = calculate_pm_kpis.calculate_all_pm_kpis(transformed_df, config)
            print(f"   [OK] Calculated KPIs")
        
        # Display KPI summary
        rca = kpi_results.get('RCA001', {})
        print(f"\n   KPI Results:")
        print(f"      Completion Rate: {rca.get('completion_rate', 0):.1f}%")
        print(f"      Target: {rca.get('target', 95):.1f}%")
        print(f"      Status: {rca.get('status', 'UNKNOWN')}")
        print(f"      Gap: {rca.get('gap', 0):.1f}%")
        print(f"      Completed: {rca.get('completed_ontime', 0)}/{rca.get('total_requiring_rca', 0)}")
        
        # Export dashboard
        print("\n6. Exporting dashboard...")
        output_path = export_pm_dashboard(kpi_results, transformed_df, output_dir='data/output')
        
        # Validation
        print("\n7. Validating export...")
        assert os.path.exists(output_path), "Output file not created"
        file_size = os.path.getsize(output_path)
        print(f"   [OK] File created: {file_size:,} bytes")
        
        print("\n" + "=" * 80)
        print("[OK] ALL TESTS PASSED!")
        print("=" * 80)
        print(f"\nDashboard location: {output_path}")
        print("Open the file to view your KPI dashboard")
        print("\nSession 4 Complete! Your KPI pipeline is ready!")
        
        return True
        
    except ImportError as e:
        print(f"\n[FAIL] Import Error: {e}")
        print("\nMake sure all previous modules are available:")
        print("   - config_loader.py (Session 1)")
        print("   - load_problem_data.py (Session 1)")
        print("   - transform_problems.py (Session 2)")
        print("   - calculate_pm_kpis.py (Session 3)")
        return False
        
    except FileNotFoundError as e:
        print(f"\n[SKIP] File Not Found: {e}")
        print("\nData files not found - this is OK for unit testing")
        print("Make sure the data files exist for full integration testing:")
        print("   - data/PYTHON_EMEA_PM_P1P2__This_Year_.csv")
        print("   - data/PYTHON_EMEA_TASK_RCA__This_Year_.csv")
        return True  # Don't fail if data files missing
        
    except Exception as e:
        print(f"\n[FAIL] Test Failed: {e}")
        import traceback
        traceback.print_exc()
        return False


# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == '__main__':
    print("\n" + "=" * 80)
    print("Problem Management KPI Dashboard Export Module")
    print("Session 4 - Excel Report Generation")
    print("=" * 80)
    
    test_export()

