"""
Problem Management KPI Dashboard Export Module

This module generates professional Excel dashboards for Problem Management KPIs.
Creates multi-sheet workbooks with KPI summaries and detailed problem breakdowns.

Author: KPI Pipeline Project - Session 4
Date: 2025-11-04
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict, Any
import os


# ============================================================================
# COLOR CONSTANTS
# ============================================================================

# Status Colors (matching traffic light system)
COLOR_GREEN = "C6EFCE"      # Light green
COLOR_YELLOW = "FFEB9C"     # Light yellow
COLOR_RED = "FFC7CE"        # Light red

# Header Colors
COLOR_HEADER = "4472C4"     # Professional blue
COLOR_WHITE = "FFFFFF"      # White text

# Border Colors
COLOR_BORDER = "D0D0D0"     # Light gray


# ============================================================================
# HELPER FUNCTIONS - FORMATTING
# ============================================================================

def get_status_color(status: str) -> str:
    """
    Get the fill color for a given status.
    
    Args:
        status: Status string ('GREEN', 'YELLOW', or 'RED')
        
    Returns:
        Hex color code for the status
        
    Example:
        >>> get_status_color('GREEN')
        'C6EFCE'
    """
    color_map = {
        'GREEN': COLOR_GREEN,
        'YELLOW': COLOR_YELLOW,
        'RED': COLOR_RED
    }
    return color_map.get(status.upper(), COLOR_BORDER)


def get_status_emoji(status: str) -> str:
    """
    Get an emoji indicator for a given status.
    
    Args:
        status: Status string ('GREEN', 'YELLOW', or 'RED')
        
    Returns:
        Emoji string for the status
        
    Example:
        >>> get_status_emoji('RED')
        '🔴'
    """
    emoji_map = {
        'GREEN': '🟢',
        'YELLOW': '🟡',
        'RED': '🔴'
    }
    return emoji_map.get(status.upper(), '⚪')


def apply_header_style(cell, bold: bool = True, bg_color: str = COLOR_HEADER):
    """
    Apply professional header styling to a cell.
    
    Args:
        cell: openpyxl cell object
        bold: Whether to make text bold
        bg_color: Background color (hex code)
    """
    cell.font = Font(bold=bold, color=COLOR_WHITE, size=11)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )
    cell.border = thin_border


def apply_cell_style(cell, bg_color: str = None, bold: bool = False, 
                     align: str = 'left', number_format: str = None):
    """
    Apply general cell styling.
    
    Args:
        cell: openpyxl cell object
        bg_color: Background color (hex code, optional)
        bold: Whether to make text bold
        align: Text alignment ('left', 'center', 'right')
        number_format: Number format string (e.g., '0.0%')
    """
    if bold:
        cell.font = Font(bold=True)
    
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    
    cell.alignment = Alignment(horizontal=align, vertical='center')
    
    if number_format:
        cell.number_format = number_format
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )
    cell.border = thin_border


def auto_adjust_column_width(worksheet, min_width: int = 12, max_width: int = 50):
    """
    Auto-adjust column widths based on content.
    
    Args:
        worksheet: openpyxl worksheet object
        min_width: Minimum column width
        max_width: Maximum column width
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


# ============================================================================
# SUMMARY SHEET CREATION
# ============================================================================

def create_summary_sheet(workbook: Workbook, kpi_results: Dict[str, Any]) -> None:
    """
    Create the KPI Summary sheet with formatted metrics and status indicators.
    
    This sheet provides a high-level overview of the RCA001 KPI including:
    - Completion rate and target
    - Status indicator with color coding
    - Gap analysis
    - Detailed counts
    - Calculation metadata
    
    Args:
        workbook: openpyxl Workbook object
        kpi_results: Dictionary with KPI calculation results from calculate_all_pm_kpis()
                    Expected keys: 'kpi_id', 'completion_rate', 'target', 'status',
                                   'gap', 'completed_ontime', 'total_requiring_rca',
                                   'total_problems', 'calculation_date'
    
    Example:
        >>> wb = Workbook()
        >>> kpi_data = calculate_all_pm_kpis(df, config)
        >>> create_summary_sheet(wb, kpi_data)
        >>> wb.save('dashboard.xlsx')
    """
    # Get or create summary sheet
    if 'Summary' in workbook.sheetnames:
        ws = workbook['Summary']
    else:
        ws = workbook.create_sheet('Summary', 0)
    
    # Extract RCA001 results
    rca_results = kpi_results.get('RCA001', {})
    
    # ========================================================================
    # HEADER SECTION
    # ========================================================================
    
    # Title
    ws['A1'] = 'Problem Management KPI Dashboard'
    ws['A1'].font = Font(bold=True, size=16, color=COLOR_HEADER)
    ws.merge_cells('A1:E1')
    
    # Subtitle with date
    calc_date = rca_results.get('calculation_date', datetime.now().strftime('%Y-%m-%d'))
    ws['A2'] = f'Report Date: {calc_date}'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')
    
    # Add spacing
    ws.row_dimensions[3].height = 5
    
    # ========================================================================
    # KPI HEADER ROW
    # ========================================================================
    
    headers = ['KPI', 'Description', 'Actual', 'Target', 'Status', 'Gap', 'Performance']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        apply_header_style(cell)
    
    # ========================================================================
    # KPI DATA ROW
    # ========================================================================
    
    row = 5
    
    # KPI ID
    ws.cell(row=row, column=1).value = rca_results.get('kpi_id', 'RCA001')
    apply_cell_style(ws.cell(row=row, column=1), bold=True)
    
    # Description
    ws.cell(row=row, column=2).value = 'Root Cause Analysis Completion Rate'
    apply_cell_style(ws.cell(row=row, column=2))
    
    # Actual (completion rate)
    actual = rca_results.get('completion_rate', 0)
    ws.cell(row=row, column=3).value = actual / 100  # Convert to decimal for percentage format
    apply_cell_style(ws.cell(row=row, column=3), align='center', number_format='0.0%')
    
    # Target
    target = rca_results.get('target', 95)
    ws.cell(row=row, column=4).value = target / 100  # Convert to decimal for percentage format
    apply_cell_style(ws.cell(row=row, column=4), align='center', number_format='0.0%')
    
    # Status (with emoji and color)
    status = rca_results.get('status', 'UNKNOWN')
    status_emoji = get_status_emoji(status)
    status_color = get_status_color(status)
    ws.cell(row=row, column=5).value = f"{status_emoji} {status}"
    apply_cell_style(ws.cell(row=row, column=5), bg_color=status_color, 
                    bold=True, align='center')
    
    # Gap
    gap = rca_results.get('gap', 0)
    ws.cell(row=row, column=6).value = gap / 100  # Convert to decimal for percentage format
    apply_cell_style(ws.cell(row=row, column=6), align='center', number_format='0.0%')
    
    # Performance (completed / total)
    completed = rca_results.get('completed_ontime', 0)
    total_rca = rca_results.get('total_requiring_rca', 0)
    ws.cell(row=row, column=7).value = f"{completed} / {total_rca}"
    apply_cell_style(ws.cell(row=row, column=7), align='center')
    
    # ========================================================================
    # DETAILS SECTION
    # ========================================================================
    
    # Add spacing
    ws.row_dimensions[6].height = 5
    
    # Details header
    ws['A7'] = 'Detailed Breakdown'
    ws['A7'].font = Font(bold=True, size=12, color=COLOR_HEADER)
    ws.merge_cells('A7:D7')
    
    # Details data
    details = [
        ('Total Problems (P1/P2):', rca_results.get('total_problems', 0)),
        ('Requiring RCA:', rca_results.get('total_requiring_rca', 0)),
        ('Completed On-Time:', rca_results.get('completed_ontime', 0)),
        ('Late or Incomplete:', total_rca - completed if total_rca > completed else 0)
    ]
    
    start_row = 8
    for idx, (label, value) in enumerate(details):
        row = start_row + idx
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
    
    # ========================================================================
    # THRESHOLD SECTION
    # ========================================================================
    
    # Add spacing
    ws.row_dimensions[12].height = 5
    
    # Thresholds header
    ws['A13'] = 'Status Thresholds'
    ws['A13'].font = Font(bold=True, size=12, color=COLOR_HEADER)
    ws.merge_cells('A13:C13')
    
    # Threshold details (these should come from config in real implementation)
    thresholds = [
        ('🟢 GREEN:', '≥ 95%'),
        ('🟡 YELLOW:', '≥ 85% and < 95%'),
        ('🔴 RED:', '< 85%')
    ]
    
    start_row = 14
    for idx, (label, criteria) in enumerate(thresholds):
        row = start_row + idx
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = criteria
    
    # ========================================================================
    # FORMATTING ADJUSTMENTS
    # ========================================================================
    
    # Auto-adjust column widths
    auto_adjust_column_width(ws, min_width=15, max_width=40)
    
    # Freeze top rows
    ws.freeze_panes = 'A5'
    
    print(f"✅ Summary sheet created successfully")


# ============================================================================
# DETAIL SHEET CREATION
# ============================================================================

def create_detail_sheet(workbook: Workbook, df: pd.DataFrame) -> None:
    """
    Create the Problem Details sheet with all RCA-requiring problems.
    
    This sheet provides a detailed view of all problems that require RCA,
    including their status, timeline, and completion information.
    
    Args:
        workbook: openpyxl Workbook object
        df: Transformed DataFrame from transform_all_problem_data()
            Must include columns: Problem_Number, Priority, State, Created_Date,
            Requires_RCA, RCA_OnTime, RCA_Stage, Days_Open, Is_Major_Problem
    
    Example:
        >>> wb = Workbook()
        >>> transformed_df = transform_all_problem_data(problems, tasks)
        >>> create_detail_sheet(wb, transformed_df)
        >>> wb.save('dashboard.xlsx')
    """
    # Get or create detail sheet
    if 'RCA Details' in workbook.sheetnames:
        ws = workbook['RCA Details']
    else:
        ws = workbook.create_sheet('RCA Details')
    
    # Filter to only problems requiring RCA
    detail_df = df[df['Requires_RCA'] == True].copy()
    
    # Select and order columns for the report
    columns_to_show = [
        'Problem_Number',
        'Priority',
        'State',
        'Created_Date',
        'Days_Open',
        'Requires_RCA',
        'RCA_OnTime',
        'RCA_Stage'
    ]
    
    # Ensure all columns exist
    available_columns = [col for col in columns_to_show if col in detail_df.columns]
    detail_df = detail_df[available_columns]
    
    # Rename columns for better readability
    column_names = {
        'Problem_Number': 'Problem Number',
        'Priority': 'Priority',
        'State': 'State',
        'Created_Date': 'Created Date',
        'Days_Open': 'Days Open',
        'Requires_RCA': 'Requires RCA',
        'RCA_OnTime': 'RCA On-Time',
        'RCA_Stage': 'RCA Stage'
    }
    detail_df = detail_df.rename(columns=column_names)
    
    # ========================================================================
    # HEADER SECTION
    # ========================================================================
    
    # Title
    ws['A1'] = 'Problem Details - RCA Required'
    ws['A1'].font = Font(bold=True, size=14, color=COLOR_HEADER)
    ws.merge_cells('A1:H1')
    
    # Subtitle
    ws['A2'] = f'Total Problems Requiring RCA: {len(detail_df)}'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:H2')
    
    # Add spacing
    ws.row_dimensions[3].height = 5
    
    # ========================================================================
    # WRITE DATA TO SHEET
    # ========================================================================
    
    # Write headers (row 4)
    for col_idx, column_name in enumerate(detail_df.columns, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = column_name
        apply_header_style(cell)
    
    # Write data (starting from row 5)
    start_row = 5
    for row_idx, row_data in enumerate(dataframe_to_rows(detail_df, index=False, header=False), start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # Apply formatting based on column
            if col_idx == 1:  # Problem Number
                apply_cell_style(cell, bold=True)
            elif col_idx == 2:  # Priority
                # Color code priorities
                if value == 'P1':
                    apply_cell_style(cell, bg_color=COLOR_RED, align='center')
                elif value == 'P2':
                    apply_cell_style(cell, bg_color=COLOR_YELLOW, align='center')
                else:
                    apply_cell_style(cell, align='center')
            elif col_idx == 5:  # Days Open
                apply_cell_style(cell, align='center')
            elif col_idx == 7:  # RCA On-Time
                # Color code RCA status
                if value == True or value == 'True':
                    apply_cell_style(cell, bg_color=COLOR_GREEN, align='center')
                elif value == False or value == 'False':
                    apply_cell_style(cell, bg_color=COLOR_RED, align='center')
                else:
                    apply_cell_style(cell, align='center')
            else:
                apply_cell_style(cell)
    
    # ========================================================================
    # FORMATTING ADJUSTMENTS
    # ========================================================================
    
    # Auto-adjust column widths
    auto_adjust_column_width(ws, min_width=12, max_width=30)
    
    # Freeze top rows and first column
    ws.freeze_panes = 'B5'
    
    print(f"✅ Detail sheet created with {len(detail_df)} problems")


# ============================================================================
# MAIN EXPORT FUNCTION
# ============================================================================

def export_pm_dashboard(kpi_results: Dict[str, Any], 
                       transformed_df: pd.DataFrame,
                       output_dir: str = 'data',
                       filename: str = None) -> str:
    """
    Create and export a complete Problem Management KPI dashboard to Excel.
    
    This is the main function that orchestrates the creation of a multi-sheet
    Excel workbook with KPI summary and problem details.
    
    Args:
        kpi_results: Dictionary with KPI calculation results from calculate_all_pm_kpis()
        transformed_df: Transformed DataFrame from transform_all_problem_data()
        output_dir: Directory to save the Excel file (default: 'data')
        filename: Optional custom filename (default: auto-generated with timestamp)
        
    Returns:
        Full path to the created Excel file
        
    Raises:
        ValueError: If required data is missing or invalid
        IOError: If unable to write the file
        
    Example:
        >>> # After calculating KPIs and transforming data
        >>> kpis = calculate_all_pm_kpis(transformed_df, config)
        >>> filepath = export_pm_dashboard(kpis, transformed_df)
        >>> print(f"Dashboard created: {filepath}")
        Dashboard created: data/PM_Dashboard_2025-11-04.xlsx
    """
    # ========================================================================
    # VALIDATION
    # ========================================================================
    
    if not kpi_results:
        raise ValueError("KPI results cannot be empty")
    
    if transformed_df is None or transformed_df.empty:
        raise ValueError("Transformed DataFrame cannot be empty")
    
    # Validate required columns
    required_columns = ['Requires_RCA', 'RCA_OnTime', 'Problem_Number']
    missing_columns = [col for col in required_columns if col not in transformed_df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")
    
    # ========================================================================
    # SETUP OUTPUT PATH
    # ========================================================================
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename if not provided
    if filename is None:
        timestamp = datetime.now().strftime('%Y-%m-%d')
        filename = f'PM_Dashboard_{timestamp}.xlsx'
    
    # Ensure .xlsx extension
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    output_path = os.path.join(output_dir, filename)
    
    # ========================================================================
    # CREATE WORKBOOK
    # ========================================================================
    
    print(f"\n📊 Creating Problem Management Dashboard...")
    print(f"   Output: {output_path}")
    
    # Create new workbook
    wb = Workbook()
    
    # Remove default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========================================================================
    # CREATE SHEETS
    # ========================================================================
    
    print(f"\n📄 Creating Summary sheet...")
    create_summary_sheet(wb, kpi_results)
    
    print(f"📄 Creating Detail sheet...")
    create_detail_sheet(wb, transformed_df)
    
    # ========================================================================
    # SAVE WORKBOOK
    # ========================================================================
    
    try:
        wb.save(output_path)
        print(f"\n✅ Dashboard created successfully!")
        print(f"   Location: {output_path}")
        print(f"   Sheets: {', '.join(wb.sheetnames)}")
        return output_path
    
    except Exception as e:
        raise IOError(f"Failed to save Excel file: {str(e)}")


# ============================================================================
# TEST FUNCTION
# ============================================================================

def test_export():
    """
    Test function to validate the export module with sample data.
    
    This function creates a sample dashboard using the EMEA Problem Management data.
    It requires that all previous modules (Sessions 1-3) are available and working.
    
    Expected output:
    - Excel file created in data/ directory
    - Summary sheet with RCA001 KPI (74.4%, RED status)
    - Detail sheet with 39 problems requiring RCA
    
    To run this test:
        python src/generate_pm_reports.py
        
    Or from Python:
        from src.generate_pm_reports import test_export
        test_export()
    """
    print("=" * 80)
    print("TESTING: Problem Management Report Export")
    print("=" * 80)
    
    try:
        # Import required modules
        print("\n1️⃣ Importing modules...")
        from config_loader import Config
        from load_problem_data import load_all_problem_data
        from transform_problems import transform_all_problem_data
        from calculate_pm_kpis import calculate_all_pm_kpis
        
        # Load configuration
        print("\n2️⃣ Loading configuration...")
        config = Config()
        print(f"   ✅ Config loaded")
        
        # Load data
        print("\n3️⃣ Loading problem data...")
        problems_df, tasks_df = load_all_problem_data('data')
        print(f"   ✅ Loaded {len(problems_df)} problems and {len(tasks_df)} tasks")
        
        # Transform data
        print("\n4️⃣ Transforming data...")
        transformed_df = transform_all_problem_data(problems_df, tasks_df)
        print(f"   ✅ Transformed {len(transformed_df)} problems")
        
        # Calculate KPIs
        print("\n5️⃣ Calculating KPIs...")
        kpi_results = calculate_all_pm_kpis(transformed_df, config)
        print(f"   ✅ Calculated KPIs")
        
        # Display KPI summary
        rca = kpi_results['RCA001']
        print(f"\n   📊 RCA001 Results:")
        print(f"      Completion Rate: {rca['completion_rate']:.1f}%")
        print(f"      Target: {rca['target']:.1f}%")
        print(f"      Status: {rca['status']}")
        print(f"      Gap: {rca['gap']:.1f}%")
        print(f"      Completed: {rca['completed_ontime']}/{rca['total_requiring_rca']}")
        
        # Export dashboard
        print("\n6️⃣ Exporting dashboard...")
        output_path = export_pm_dashboard(kpi_results, transformed_df)
        
        # Validation
        print("\n7️⃣ Validating export...")
        assert os.path.exists(output_path), "Output file not created"
        file_size = os.path.getsize(output_path)
        print(f"   ✅ File created: {file_size:,} bytes")
        
        print("\n" + "=" * 80)
        print("✅ ALL TESTS PASSED!")
        print("=" * 80)
        print(f"\n📁 Dashboard location: {output_path}")
        print(f"📊 Open the file to view your KPI dashboard")
        print("\n✨ Session 4 Complete! Your KPI pipeline is ready! ✨")
        
        return True
        
    except ImportError as e:
        print(f"\n❌ Import Error: {e}")
        print("\n💡 Make sure all previous modules are in the same directory:")
        print("   - config_loader.py (Session 1)")
        print("   - load_problem_data.py (Session 1)")
        print("   - transform_problems.py (Session 2)")
        print("   - calculate_pm_kpis.py (Session 3)")
        return False
        
    except FileNotFoundError as e:
        print(f"\n❌ File Not Found: {e}")
        print("\n💡 Make sure data files are in the 'data/' directory:")
        print("   - PYTHON_EMEA_PM_P1P2__This_Year_.csv")
        print("   - PYTHON_EMEA_PM_RCA_Tasks__This_Year_.csv")
        return False
        
    except Exception as e:
        print(f"\n❌ Test Failed: {e}")
        import traceback
        traceback.print_exc()
        return False


# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == '__main__':
    print("\n" + "🚀" * 40)
    print("Problem Management KPI Dashboard Export Module")
    print("Session 4 - Excel Report Generation")
    print("🚀" * 40)
    
    test_export()
